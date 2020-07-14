from PyQt5.QtWidgets import QApplication, QMainWindow,QGridLayout,QPushButton
from PyQt5.QtCore import QSize, Qt,QEvent
from PyQt5.QtWidgets import QWidget,QTableWidget,QTableWidgetItem,QFormLayout,QLineEdit
from PyQt5.QtGui import QIcon  
from PyQt5 import QtCore
from openpyxl import load_workbook
import os                
import sys
import sqlite3
database="files/effectif.db"
excel_path="files/table.xlsx"	
class e_s(QWidget):
    def __init__(self):
        super().__init__()
        QWidget.__init__(self)
        self.setGeometry(100,50,1200,600)
        self.setMaximumSize(1200,600)
        self.setMinimumSize(1200,600)
        self.setWindowTitle("ENTREE SORTIE")
        self.setWindowIcon(QIcon("icons/in_out.png"))
        #table parametre
        self.table=QTableWidget(self)
        self.table.setColumnCount(6)
        self.table.setRowCount(10)
        self.table.setGeometry(10,60,1000,510)
        self.table.setColumnWidth(0,200)
        self.table.setColumnWidth(1,150)
        self.table.setColumnWidth(2,150)
        self.table.setColumnWidth(3,150)
        self.table.setColumnWidth(4,150)
        self.table.setColumnWidth(5,150)
        self.update_table()
        #loading icons 
        del_icon=QIcon("icons/delet1.png")
        print_icon=QIcon("icons/print_icon.png")
        update_icon=QIcon("icons/update2.png")
        format_icon=QIcon("icons/format.png")
        #creating filters
        self.name_filtre=QLineEdit(self)
        self.mat_filtre=QLineEdit(self)
        self.code_filtre=QLineEdit(self)
        self.date_filtre=QLineEdit(self)
        self.dir_filtre=QLineEdit(self)
        self.heure_filtre=QLineEdit(self)
        #adding filtre to window
        self.name_filtre.setGeometry(10,35,225,20)
        self.mat_filtre.setGeometry(240,35,145,20)
        self.code_filtre.setGeometry(390,35,148,20)
        self.date_filtre.setGeometry(543,35,145,20)
        self.dir_filtre.setGeometry(690,35,148,20)
        self.heure_filtre.setGeometry(840,35,150,20)
        #addin_event
        self.name_filtre.installEventFilter(self)
        self.mat_filtre.installEventFilter(self)
        self.code_filtre.installEventFilter(self)
        self.date_filtre.installEventFilter(self)
        self.dir_filtre.installEventFilter(self)
        self.heure_filtre.installEventFilter(self)
        #adding buttons
        self.btn_print=QPushButton("Imprimer",self)
        self.btn_print.setGeometry(1030,60,140,60)
        self.btn_delete=QPushButton("Supprimer",self)
        self.btn_delete.setGeometry(1030,130,140,60)
        self.btn_update=QPushButton("Mettre a jour",self)
        self.btn_update.setGeometry(1030,200,140,60)
        self.btn_format=QPushButton("Formater",self)
        self.btn_format.setGeometry(1030,270,140,60)
        #buttons icons
        self.btn_print.setIcon(print_icon)
        self.btn_delete.setIcon(del_icon)
        self.btn_update.setIcon(update_icon)
        self.btn_format.setIcon(format_icon)
        #button functions
        self.btn_format.clicked.connect(self.format_table)
        self.btn_update.clicked.connect(self.update_table)
        self.btn_print.clicked.connect(self.print_table)
        self.btn_delete.clicked.connect(self.delete_selected_items)
        #icons size
        self.btn_print.setIconSize(QSize(50,50))
        self.btn_delete.setIconSize(QSize(50,50))
        self.btn_update.setIconSize(QSize(50,50))
        self.btn_format.setIconSize(QSize(50,50))
        #button styling
        self.btn_delete.setObjectName("btn1")
        self.btn_print.setObjectName("btn1")
        self.btn_update.setObjectName("btn1")
        self.btn_format.setObjectName("btn1")

    def update_table(self):
        conn=sqlite3.connect(database)
        c=conn.cursor()
        c.execute("SELECT *FROM entre_sortie")
        conn.commit()
        data=c.fetchall()
        self.table.clear()
        self.table.setColumnWidth(0,225)
        self.table.setColumnWidth(1,153)
        self.table.setColumnWidth(2,150)
        self.table.setColumnWidth(3,150)
        self.table.setColumnWidth(4,150)
        self.table.setColumnWidth(5,150)
        self.table.setHorizontalHeaderLabels(["Nom et Prenom", "Matricule","Code","Date","Direction","Heure"])
        self.table.setRowCount(len(data))
        for i in range(len(data)):
            st=data[i]
            for j in range(len(st)):
                self.table.setItem(i,j,QTableWidgetItem(st[j]))
        conn.close()
    
    def filtre_table(self,name,mat,code,datee,dir,heure):
        name=name+'%'
        mat=mat+'%'
        code=code+'%'
        datee=datee+'%'
        dir=dir+'%'
        heure=heure+'%'
        sqstring="SELECT *FROM entre_sortie WHERE name LIKE '{}' AND matricule LIKE '{}' AND code LIKE '{}' AND _date LIKE '{}' AND uveme LIKE '{}' AND temps LIKE '{}'"
        self.table.clear()
        self.table.setColumnWidth(0,200)
        self.table.setColumnWidth(1,150)
        self.table.setColumnWidth(2,150)
        self.table.setColumnWidth(3,150)
        self.table.setColumnWidth(4,150)
        self.table.setColumnWidth(5,150)
        self.table.setHorizontalHeaderLabels(["Nom et Prenom", "Matricule","Code","Date","Direction","Heure"])
        conn=sqlite3.connect(database)
        c=conn.cursor()
        c.execute(sqstring.format(name,mat,code,datee,dir,heure))
        conn.commit()
        data=c.fetchall()
        self.table.setRowCount(len(data))
        for i in range(len(data)):
            st=data[i]
            for j in range(len(st)):
                self.table.setItem(i,j,QTableWidgetItem(st[j]))
        conn.close()
    def format_table(self):
        conn=sqlite3.connect(database)
        c=conn.cursor()
        c.execute("DELETE FROM entre_sortie")
        conn.commit()
        #data=c.fetchall()
        conn.close()
        self.update_table()

    def delete_selected_items(self):
        items=self.table.selectedItems()
        if (len(items)%6==0):
            conn=sqlite3.connect(database)
            c=conn.cursor()
            for i in range(int(len(items)/6)):
                #print(i+1)
                st=items[i*6:(i*6)+6]
                c.execute("DELETE FROM entre_sortie WHERE code='{}' AND uveme='{}' AND temps='{}'".format(st[2].text(),st[4].text(),st[5].text()))
                conn.commit()
                """
                for j in range(len(st)):
                    print("index",j,st[j].text())
                """
            conn.close()
            self.update_table()

    def print_table(self):
        wb=load_workbook(excel_path)
        std=wb.get_sheet_by_name('Feuil1')
        wb.remove_sheet(std)
        s=wb.create_sheet('Feuil1')
        for i in range(self.table.rowCount()):
            list1=[self.table.item(i,0).text(),self.table.item(i,1).text(),self.table.item(i,2).text(),self.table.item(i,3).text(),self.table.item(i,4).text(),self.table.item(i,5).text()]
            s.append(list1)
        wb.save(excel_path)
        os.system("start excel files/table ")
    def eventFilter(self,source,event):
        if event.type()==QEvent.KeyPress:
            name=self.name_filtre.text()
            mat=self.mat_filtre.text()
            code=self.code_filtre.text()
            datee=self.date_filtre.text()
            dir=self.dir_filtre.text()
            heure=self.heure_filtre.text()
            if event.key()!=16777219:
                if source is self.name_filtre:
                    name=name+event.text()
                if source is self.code_filtre:
                    code=code+event.text()
                if source is self.mat_filtre:
                    mat=mat+event.text()
                if source is self.date_filtre:
                    datee=datee+event.text()
                if source is self.dir_filtre:
                    dir=dir+event.text()
                if source is self.heure_filtre:
                    heure=heure+event.text()
            if event.key()==16777219:
                if source is self.name_filtre:
                    name=name[0:len(name)-1]
                if source is self.code_filtre:
                    code=code[0:len(code)-1]
                if source is self.mat_filtre:
                    mat=mat[0:len(mat)-1]
                if source is self.date_filtre:
                    datee=datee[0:len(datee)-1]
                if source is self.dir_filtre:
                    dir=dir[0:len(dir)-1]
                if source is self.heure_filtre:
                    heure=heure[0:len(heure)-1]
            self.filtre_table(name,mat,code,datee,dir,heure)
        return super(e_s, self).eventFilter(source, event)