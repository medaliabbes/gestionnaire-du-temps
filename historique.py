import sys 
import sqlite3  
import os
from openpyxl import load_workbook
from PyQt5.QtCore import QSize,QEvent
from PyQt5.QtGui import *
from PyQt5 import QtWidgets
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import QApplication,QMainWindow,QPushButton,QWidget,QLineEdit,QTableWidget,QTableWidgetItem
database="files/effectif.db"
excel_path="files/table.xlsx"
class history(QWidget):
    def __init__(self):
        super().__init__()
        QWidget.__init__(self)
        self.setGeometry(100,50,1200,600)
        self.setMaximumSize(1200,600)
        self.setMinimumSize(1200,600)
        self.setWindowTitle("HISTORIQUE")
        self.setWindowIcon(QIcon("icons/history.png"))
        self.table=QTableWidget(self)
        self.table.setColumnCount(7)
        self.table.setRowCount(10)
        self.table.setGeometry(10,60,1010,510)
        self.table.setHorizontalHeaderLabels(["Nom et Prenom", "Code", "Matricule","date","sortie","retour","durée"])
        self.table.setColumnWidth(0,260)
        self.table.setColumnWidth(3,150)
        self.table.setColumnWidth(2,150)
        self.update_tab()
        #icons
        del_icon=QIcon("icons/delet1.png")
        print_icon=QIcon("icons/print_icon.png")
        update_icon=QIcon("icons/update2.png")
        format_icon=QIcon("icons/format.png")
        #button_imprimer
        self.btn_print=QPushButton("imprimer",self)
        self.btn_print.setGeometry(1030,60,150,60)
        self.btn_print.setIcon(print_icon)
        self.btn_print.setIconSize(QSize(50,50))
        self.btn_print.clicked.connect(self.print_table)
        #button_suprimer
        self.btn_del=QPushButton("supprimer",self)
        self.btn_del.setGeometry(1030,130,150,60)
        self.btn_del.setIcon(del_icon)
        self.btn_del.setIconSize(QSize(50,50))
        self.btn_del.clicked.connect(self.delete_selected_items)
        #button_update
        self.btn_update=QPushButton("mettre a jour",self)
        self.btn_update.setGeometry(1030,200,150,60)
        self.btn_update.setIcon(update_icon)
        self.btn_update.setIconSize(QSize(50,50))
        self.btn_update.clicked.connect(self.update_tab)
        #button_format
        self.btn_format=QPushButton("formater",self)
        self.btn_format.setGeometry(1030,270,150,60)
        self.btn_format.setIcon(format_icon)
        self.btn_format.setIconSize(QSize(50,50))
        self.btn_format.clicked.connect(self.format_table)
        #button style
        self.btn_del.setObjectName("btn1")
        self.btn_format.setObjectName("btn1")
        self.btn_print.setObjectName("btn1")
        self.btn_update.setObjectName("btn1")
        #entry_fields#filtre
        self.filtre_name=QLineEdit(self)
        self.filtre_name.setGeometry(10,33,278,20)

        self.code_filtre=QLineEdit(self)
        self.code_filtre.setGeometry(295,33,95,20)

        self.mat_filtre=QLineEdit(self)
        self.mat_filtre.setGeometry(393,33,148,20)

        self.date_filtre=QLineEdit(self)
        self.date_filtre.setGeometry(545,33,145,20)

        self.sortie_filtre=QLineEdit(self)
        self.sortie_filtre.setGeometry(695,33,95,20)

        self.retour_filtre=QLineEdit(self)
        self.retour_filtre.setGeometry(795,33,95,20)

        self.duree_filtre=QLineEdit(self)
        self.duree_filtre.setGeometry(895,33,95,20)
        #filtre_event
        self.filtre_name.installEventFilter(self)
        self.code_filtre.installEventFilter(self)
        self.mat_filtre.installEventFilter(self)
        self.date_filtre.installEventFilter(self)
        self.sortie_filtre.installEventFilter(self)
        self.retour_filtre.installEventFilter(self)
        self.duree_filtre.installEventFilter(self)

    def update_tab(self):
        conn=sqlite3.connect(database)
        c=conn.cursor()
        c.execute("SELECT *FROM historiqueES ")
        conn.commit()
        data=c.fetchall()
        self.table.clear()
        self.table.setRowCount(len(data))
        self.table.setHorizontalHeaderLabels(["Nom et Prenom", "Code", "Matricule","date","sortie","retour","durée"])
        for i in range(len(data)):
            st=data[i]
            for j in range(len(st)):
                self.table.setItem(i,j, QTableWidgetItem(st[j]))
    def delete_selected_items(self):
        items=self.table.selectedItems()
        conn=sqlite3.connect(database)
        c=conn.cursor()
        if (len(items)%7==0) :
            for i in range(int(len(items)/7)):
                data=items[i*7:i*7+7]
                c.execute("DELETE FROM historiqueES WHERE code='{}' AND sortie='{}' AND retour='{}'".format(data[1].text(),data[4].text(),data[5].text()))
                conn.commit()
                
            conn.close()
        else :
            return
        self.update_tab()
    def format_table(self):
        conn=sqlite3.connect(database)
        c=conn.cursor()
        c.execute("DELETE FROM historiqueES")
        conn.commit()
        conn.close()
        self.update_tab()
    def filtre_table(self,name,code,mat,datee,sortie,retour,duree):
        name=name+'%'
        code=code+'%'
        mat=mat+'%'
        datee=datee+'%'
        sortie=sortie+'%'
        retour=retour+'%'
        duree=duree+'%'
        conn=sqlite3.connect(database)
        c=conn.cursor()
        sqlstring="SELECT *FROM historiqueES WHERE nom LIKE '{}' AND code LIKE '{}' AND matricule LIKE '{}' AND _date LIKE '{}' AND sortie LIKE '{}' AND retour LIKE '{}' AND dure LIKE '{}'"
        c.execute(sqlstring.format(name,code,mat,datee,retour,sortie,duree))
        conn.commit()
        data=c.fetchall()
        conn.close()
        self.table.clear()
        size_tab=len(data)
        self.table.setRowCount(size_tab)
        self.table.setHorizontalHeaderLabels(["Nom et Prenom", "Code", "Matricule","date","sortie","retour","durée"])
        for i in range(size_tab):
            st=data[i]
            for j in range(len(st)):
                self.table.setItem(i,j, QTableWidgetItem(st[j]))
    def print_table(self):
        wb=load_workbook(excel_path)
        std=wb.get_sheet_by_name('Feuil1')
        wb.remove_sheet(std)
        s=wb.create_sheet('Feuil1')
        for i in range(self.table.rowCount()):
            list1=[self.table.item(i,0).text(),self.table.item(i,1).text(),self.table.item(i,2).text(),self.table.item(i,3).text(),self.table.item(i,4).text(),self.table.item(i,5).text(),self.table.item(i,6).text()]
            s.append(list1)
        wb.save(excel_path)
        os.system("start excel files/table ")
    def eventFilter(self,source,event):
        if (event.type() == QEvent.KeyPress ):
            name=self.filtre_name.text()
            code=self.code_filtre.text()
            mat=self.mat_filtre.text()
            sortie=self.sortie_filtre.text()
            retour=self.retour_filtre.text()
            datee=self.date_filtre.text()
            duree=self.duree_filtre.text()
            if event.key()!=16777219:
                if source is self.filtre_name:
                    name=name+event.text()
                if source is self.code_filtre:
                    code=code+event.text()
                if source is self.mat_filtre:
                    mat=mat+event.text()
                if source is self.date_filtre:
                    datee=datee+event.text()
                if source is self.sortie_filtre:
                    sortie=sortie+event.text()
                if source is self.retour_filtre:
                    retour=retour+event.text()
                if source is self.duree_filtre:
                    duree=duree+event.text()
            if event.key()==16777219:
                if source is self.filtre_name:
                    name=name[0:len(name)-1]
                if source is self.code_filtre:
                    code=code[0:len(code)-1]
                if source is self.mat_filtre:
                    mat=mat[0:len(mat)-1]
                if source is self.date_filtre:
                    datee=datee[0:len(datee)-1]
                if source is self.sortie_filtre:
                    sortie=sortie[0:len(sortie)-1]
                if source is self.retour_filtre:
                    retour=retour[0:len(retour)-1]
                if source is self.duree_filtre:
                    duree=duree[0:len(duree)-1]
            self.filtre_table(name,code,mat,datee,sortie,retour,duree)
        return super(history, self).eventFilter(source, event)
