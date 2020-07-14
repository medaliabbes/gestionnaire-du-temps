from PyQt5.QtWidgets import QApplication, QMainWindow,QGridLayout,QPushButton
from PyQt5.QtCore import QSize, Qt,QEvent
from PyQt5.QtWidgets import QWidget,QTableWidget,QTableWidgetItem,QFormLayout,QLineEdit,QLabel
from PyQt5.QtGui import QIcon  
from PyQt5 import QtCore
from openpyxl import load_workbook 
from new_operateur import user
import os                
import sys
import sqlite3

database='files/effectif.db'
excel_path="files/table.xlsx"

class users(QWidget):
	#global filt1,filt2,filt3,filt4
	global tabSize
	tabSize=0
	def __init__(self):
		global tab
		super().__init__()
		QWidget.__init__(self)
		self.setMaximumSize(1000,600) 
		self.setGeometry(250,50,1000,600)        
		self.setWindowTitle("OPERATEURS") 
		#Icons
		delete_icon=QIcon("icons/delet1.png") 
		add_icon=QIcon("icons/n_user.png") 
		print_icon=QIcon("icons/print_icon.png")  
		window_icon=QIcon("icons/users_icon.png")
		update_icon=QIcon("icons/update2.png")
		#####  
		self.setWindowIcon(window_icon) 
		tab= QTableWidget(self) 
		tab.setColumnCount(4)
		tabSize=self.table_size()
		tab.setRowCount(tabSize)
        #tab.setRowCount(100)
		tab.setHorizontalHeaderLabels(["Nom et Prenom", "Code", "Matricule","Fonction"])
		tab.setColumnWidth(0,300)
		tab.setColumnWidth(1,150)
		tab.setColumnWidth(2,200)
		self.update_table()
		tab.setGeometry(10,60,800,500)
		#filtre
		self.name_filtre=QLineEdit(self)
		self.name_filtre.setGeometry(10,29,326,20)
		self.code_filtre=QLineEdit(self)
		self.code_filtre.setGeometry(340,29,147,20)
		self.mat_filtre=QLineEdit(self)
		self.mat_filtre.setGeometry(493,29,193,20)
		self.fonc_filtre=QLineEdit(self)
		self.fonc_filtre.setGeometry(693,29,110,20)
		#enable_keypressed_for_filtre
		#self.name_filtre.setFocusPolicy(Qt.StrongFocus )
		self.name_filtre.installEventFilter(self)
		self.code_filtre.installEventFilter(self)
		self.mat_filtre.installEventFilter(self)
		self.fonc_filtre.installEventFilter(self)
		#self.name_filtre.textEdited.connect(self.showCurrentText)
		#self.code_filtre.setFocusPolicy(Qt.StrongFocus )
		#self.mat_filtre.setFocusPolicy(Qt.StrongFocus )
		#self.fonc_filtre.setFocusPolicy(Qt.StrongFocus )
		#button_ajouter
		self.h=20
		self.x=10
		btn_add=QPushButton("ajouter",self)
		btn_add.setGeometry(840,60,120+self.h,60)
		btn_add.clicked.connect(self.add)
		btn_add.setIcon(add_icon)
		btn_add.setIconSize(QSize(50,50))
		self.dia=user()#new_user_fen()
		#button_supprimer
		btn_del=QPushButton("supprimer",self)
		btn_del.setGeometry(840,130,120+self.h,60)
		btn_del.clicked.connect(self.selected_items_to_delet)
		btn_del.setIcon(delete_icon)
		btn_del.setIconSize(QSize(50,50))
		#button_imprimer
		btn_print=QPushButton("imprimer",self)
		btn_print.setGeometry(840,200,120+self.h,60)
		btn_print.setIcon(print_icon)
		btn_print.clicked.connect(self.print_table)
		btn_print.setIconSize(QSize(50,50))
		#button_update
		btn_update=QPushButton("mettre a jour",self)
		btn_update.setGeometry(840,270,120+self.h,60)
		btn_update.clicked.connect(self.update_table)
		btn_update.setIcon(update_icon)
		btn_update.setIconSize(QSize(50,50))
		#adding user from an excel file
		self.label1=QLabel("Path du fichier :",self)
		self.label1.setGeometry(10,570,90,20)
		self.path_entry=QLineEdit(self)
		self.path_entry.setGeometry(100,570,250,20)
		self.label2=QLabel("Nombre :",self)
		self.label2.setGeometry(360,570,80,20)
		self.nbr_entry=QLineEdit(self)
		self.nbr_entry.setGeometry(420,570,50,20)
		#adding the read btn
		self.btn_getPath=QPushButton("ajouter",self)
		self.btn_getPath.setGeometry(490,570,150,20)
		self.btn_getPath.clicked.connect(self.get_exceldata)
		self.btn_getPath.setObjectName("btn1")

		btn_add.setObjectName("btn1")
		btn_del.setObjectName("btn1")
		btn_print.setObjectName("btn1")
		btn_update.setObjectName("btn1")
	def get_exceldata(self):
		path=self.path_entry.text()
		#2print(path)
		path1=""
		for i in path:
			path1=path1+i
			if i==chr(92):
				path1=path1+i
		#print(path1)
		try:
		    nbr=int(self.nbr_entry.text())
		    #print(nbr)
		except :
			#print("error :",sys.exc_info())
			nbr=50
		#☺print(nbr)
		try :
			conn=sqlite3.connect(database)
			c=conn.cursor()
			path=path+".xlsx"
			excel=load_workbook(filename =path )
			sheet_ranges = excel['Feuil1']
			for i in range(nbr):
				col1='A'+str(i+1)
				col2='B'+str(i+1)
				col3='C'+str(i+1)
				col4='D'+str(i+1)
				if sheet_ranges[col2].value==None:
					break
				#str1=sheet_ranges[col1].value+' $ '+sheet_ranges[col2].value+' $ '+sheet_ranges[col3].value
				#print(str1)
				c.execute("INSERT INTO users VALUES('{}','{}','{}','{}')".format(sheet_ranges[col1].value,sheet_ranges[col2].value,sheet_ranges[col3].value,sheet_ranges[col4].value))
				conn.commit()
			conn.close()
		except :
			print("error ",sys.exc_info())
			return
		self.update_table()
		

		
	def fill_table(self,data):
		tab.clear()
		tab.setHorizontalHeaderLabels(["Nom et Prenom", "Code", "Matricule","Fonction"])
		for i in range(len(data)):
			st=data[i]
			for j in range(len(st)):
				tab.setItem(i,j, QTableWidgetItem(st[j]))
	def selected_items_to_delet(self):
		items=tab.selectedItems() 
		#print(len(items))
		if (len(items))%4==0:
			conn=sqlite3.connect(database)
			c=conn.cursor()
			for i in range(len(items)):
				if i%4==0:
					c.execute("DELETE FROM users WHERE code='{}' ".format(items[i+1].text()))
					conn.commit()
					#print(items[i+1].text())
			conn.close() 
		self.update_table()
		
		

	def filtre_tab(self,fil1,fil2,fil3,fil4):
		fil1=fil1+'%'
		fil2=fil2+'%'
		fil3=fil3+'%'
		fil4=fil4+'%'
		conn=sqlite3.connect(database)
		c=conn.cursor()
		c.execute("SELECT *FROM users WHERE name LIKE '{}' AND code LIKE '{}' AND matricule LIKE '{}' AND fonction LIKE '{}'".format(fil1,fil2,fil3,fil4))
		conn.commit()
		data=c.fetchall()
		tabSize=len(data)
		tab.setRowCount(tabSize)
		self.fill_table(data)
		conn.close()	
	def add(self):
		self.dia.show()
	def print_table(self):
		wb=load_workbook(excel_path)
		std=wb.get_sheet_by_name('Feuil1')
		wb.remove_sheet(std)
		s=wb.create_sheet('Feuil1')
		for i in range(tab.rowCount()):
			l=[tab.item(i,0).text(),tab.item(i,1).text(),tab.item(i,2).text(),tab.item(i,3).text()]
			s.append(l)
		wb.save(excel_path)
		os.system("start excel files/table ")
	def delet_user(self):
		row=tab.currentRow()
		column=tab.currentColumn()
		#item=tab.currentItem()
		#print("column :'{}'/row : '{}'".format(column,row))
		#print(tab.item(row,1).text())
		conn=sqlite3.connect(database)
		c=conn.cursor()
		c.execute("DELETE FROM users WHERE code='{}'".format(tab.item(row,1).text()))
		conn.commit()
		conn.close()
		self.update_table()
	"""
	def add_new_user(self):
		conn=sqlite3.connect(database)
		c=conn.cursor()
		conn.close()
	"""
	def update_table(self):
		conn=sqlite3.connect(database)
		c=conn.cursor()
		c.execute("SELECT *FROM users ")
		conn.commit()
		data=c.fetchall()
		conn.close()
		tab.clear()
		tab.setHorizontalHeaderLabels(["Nom et Prenom", "Code", "Matricule","Fonction"])
		tab.setRowCount(len(data))
		for i in range(len(data)):
			st=data[i]
			for j in range(len(st)):
				tab.setItem(i,j, QTableWidgetItem(st[j]))
	def table_size(self):
		conn=sqlite3.connect(database)
		c=conn.cursor()
		c.execute("SELECT *FROM users")
		conn.commit()
		d=c.fetchall()
		conn.close()
		return len(d)
	def eventFilter(self, source, event):
		if (event.type() == QEvent.KeyPress ):#and source is self.name_filtre or source is self.code_filtre or source is self.mat_filtre):
			#print('key press:', (event.key(), event.text()),self.name_filtre.text())
			#print('name:',self.name_filtre.text()+event.text())
			filt1=self.name_filtre.text()
			filt2=self.code_filtre.text()
			filt3=self.mat_filtre.text()
			filt4=self.fonc_filtre.text()
			if event.key()!=16777219:
				if source is self.name_filtre:
					filt1=filt1+event.text()
				if source is self.code_filtre:
					filt2=filt2+event.text()
				if source is self.mat_filtre:
					filt3=filt3+event.text()
				if source is self.fonc_filtre:
					filt4=filt4+event.text()
			if event.key()==16777219:
				if source is self.name_filtre:
					filt1=filt1[0:len(filt1)-1]
				if source is self.code_filtre:
					filt2=filt2[0:len(filt2)-1]
				if source is self.mat_filtre:
					filt3=filt3[0:len(filt3)-1]
				if source is self.fonc_filtre:
					filt4=filt4[0:len(filt4)-1]
			self.filtre_tab(filt1,filt2,filt3,filt4)
			tabSize=tab.rowCount()
			#print(tabSize)
		return super(users, self).eventFilter(source, event)